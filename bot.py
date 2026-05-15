"""
FinBot — Asistente financiero personal para Telegram
Versión 3.0 — Código limpio, sin bugs
"""

import os
import io
import re
import base64
import asyncio
import smtplib
import psycopg2
import psycopg2.extras
import psycopg2.pool
import openpyxl
import matplotlib
import datetime
import threading
import requests

matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl.styles import Font, PatternFill, Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from http.server import HTTPServer, BaseHTTPRequestHandler
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, MessageHandler, CommandHandler,
    CallbackQueryHandler, filters, ContextTypes,
)

# ── Configuración ────────────────────────────────────────────────────────────

TELEGRAM_TOKEN    = os.environ.get("TELEGRAM_TOKEN")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
DATABASE_URL      = os.environ.get("DATABASE_URL")
EMAIL_SENDER      = os.environ.get("EMAIL_SENDER", "")
EMAIL_PASSWORD    = os.environ.get("EMAIL_PASSWORD", "")
SMTP_SERVER       = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT         = int(os.environ.get("SMTP_PORT", "587"))

SYSTEM_PROMPT = """Eres FinBot, asistente financiero personal en Telegram. Eres amable, directo y usas emojis con criterio.

REGLAS DE FORMATO — CRITICO:
- NUNCA uses ** negritas **, tablas con |---| ni markdown de ningun tipo.
- Solo texto plano con emojis. El token SIEMPRE va en la PRIMERA linea, sin nada antes.
- Despues del token puedes dar contexto util (hasta 4 oraciones). Se especifico con los datos del usuario.

TOKENS EXACTOS (primera linea, sin espacios extra):
  GASTO_REGISTRADO|monto|categoria|descripcion
  INGRESO_REGISTRADO|monto|descripcion
  PRESUPUESTO_DEFINIDO|categoria|monto
  EMAIL_CONFIGURADO|correo
  SALDO_ACTUALIZADO|monto
  DEUDA_REGISTRADA|tipo|monto|persona|descripcion
  GASTO_FIJO_REGISTRADO|descripcion|monto|categoria
  META_REGISTRADA|nombre|monto_objetivo
  BORRAR_ULTIMO_GASTO

Categorias validas: Comida, Transporte, Compras, Salud, Entretenimiento, Servicios, Deudas, Ahorro, General

EJEMPLOS CORRECTOS:
  Usuario: "gaste 45000 en almuerzo y taxi"
  Respuesta:
  GASTO_REGISTRADO|45000|Comida|almuerzo y taxi
  Listo! Eso equivale a unos 3 almuerzos si lo miras mensualmente. Considera llevar algo de casa 2 veces por semana para ahorrar 🍱

  Usuario: "gane 3500000 de salario"
  Respuesta:
  INGRESO_REGISTRADO|3500000|Salario
  Excelente! Con la regla 50/30/20 te recomiendo reservar $700.000 para ahorro desde hoy. Primero pagarte a ti mismo 💪

  Usuario: "quiero ahorrar 2000000 para vacaciones"
  Respuesta:
  META_REGISTRADA|Vacaciones|2000000
  Meta creada! Si ahorras $167.000 al mes llegas en un ano. Puedes abonar diciendome "abone 100000 a vacaciones" 🎯

  Usuario: "debo 80000 a Carlos por el taxi"
  Respuesta:
  DEUDA_REGISTRADA|debo|80000|Carlos|taxi
  Deuda con Carlos anotada. Te aviso cuando la registres como pagada para mantener tus cuentas al dia 👍

  Usuario: "me equivoque" / "borra el ultimo"
  Respuesta:
  BORRAR_ULTIMO_GASTO
  Listo, revertido!

Para preguntas generales responde en espanol, con profundidad util, sin markdown. Si el usuario pregunta por su situacion financiera o pide un analisis, usa los datos que tienes en el historial para dar respuestas concretas y personalizadas."""

TIPS = [
    "Regla 50/30/20: 50% necesidades, 30% gustos, 20% ahorro 💡",
    "Los microgastos diarios suman mas de lo que crees. Anotatlos todos 📌",
    "Antes de comprar algo: necesidad o deseo? 🤔",
    "Un fondo de emergencia de 3-6 meses de gastos es tu mejor seguro 🛡️",
    "Ahorra primero, gasta despues. Pegate a ti mismo primero 💰",
    "Revisa tus suscripciones activas. Seguro hay dinero escondido ahi 🔍",
    "Compara precios antes de comprar grande. 10 min pueden ahorrarte mucho ⏱️",
    "Espera 24h antes de una compra impulsiva. Casi siempre lo agradeces 😌",
    "Llevar registro de deudas evita malentendidos y dinero perdido 📋",
    "Un presupuesto no te limita, te da libertad para gastar sin culpa ✅",
]

COLORES = ["#e94560","#06d6a0","#ffd166","#118ab2","#533483","#ef476f","#ffb703","#26547c","#073b4c","#0f3460"]

# ── Helpers generales ────────────────────────────────────────────────────────

def colombia_now():
    return datetime.datetime.utcnow() - datetime.timedelta(hours=5)

def colombia_today():
    return colombia_now().date()

def barra(actual, maximo, largo=10):
    if maximo <= 0:
        return "░" * largo
    pct = min(actual / maximo, 1.0)
    return "█" * int(pct * largo) + "░" * (largo - int(pct * largo))

def frase_gasto(monto):
    if monto <= 5000:    return "Microgasto anotado 📌"
    if monto <= 20000:   return "Listo, registrado ✅"
    if monto <= 100000:  return "Anotado. Cada peso cuenta 💡"
    if monto <= 500000:  return "Gasto importante registrado 📊"
    return "Gasto grande. Revisa tu presupuesto 👀"

def frase_ingreso(monto):
    if monto <= 50000:    return "Ingresito anotado 💵"
    if monto <= 500000:   return "Buen ingreso registrado 💰"
    if monto <= 2000000:  return "Muy bien! Ese ingreso suma mucho 🙌"
    return "Excelente ingreso! A administrarlo bien 🚀"

def _limpiar_markdown(texto):
    texto = re.sub(r'\*\*(.+?)\*\*', r'\1', texto)
    texto = re.sub(r'\*(.+?)\*',     r'\1', texto)
    texto = re.sub(r'\|[-:|]+\|',    '',    texto)
    texto = re.sub(r'\s*\|',         '',    texto)
    return texto.strip()

def _parse_monto(s):
    s = s.replace("$","").replace(" ","").strip()
    if s.count(".") > 1:
        s = s.replace(".","")
    return float(s.replace(",",""))

# ── Base de datos ────────────────────────────────────────────────────────────

_db_pool = None

def get_pool():
    global _db_pool
    if _db_pool is None:
        _db_pool = psycopg2.pool.ThreadedConnectionPool(
            minconn=2, maxconn=10, dsn=DATABASE_URL, sslmode="require"
        )
    return _db_pool

def get_conn():
    return get_pool().getconn()

def release_conn(conn):
    try:
        get_pool().putconn(conn)
    except Exception:
        pass

def _db_execute_raw(sql_list, commit=True):
    """Ejecuta una lista de SQL en una sola conexion del pool."""
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            for item in sql_list:
                if isinstance(item, tuple):
                    cur.execute(item[0], item[1])
                else:
                    cur.execute(item)
        if commit:
            conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        release_conn(conn)

def _db_query(sql, params=(), cursor_factory=None):
    """Ejecuta un SELECT y devuelve todos los resultados."""
    conn = get_conn()
    try:
        kwargs = {"cursor_factory": cursor_factory} if cursor_factory else {}
        with conn.cursor(**kwargs) as cur:
            cur.execute(sql, params)
            return cur.fetchall()
    finally:
        release_conn(conn)

def _db_query_one(sql, params=(), cursor_factory=None):
    """Ejecuta un SELECT y devuelve una fila."""
    conn = get_conn()
    try:
        kwargs = {"cursor_factory": cursor_factory} if cursor_factory else {}
        with conn.cursor(**kwargs) as cur:
            cur.execute(sql, params)
            return cur.fetchone()
    finally:
        release_conn(conn)

def init_db():
    tablas = [
        """CREATE TABLE IF NOT EXISTS gastos (
            id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
            monto NUMERIC NOT NULL, categoria TEXT NOT NULL,
            descripcion TEXT NOT NULL, fecha DATE NOT NULL)""",
        """CREATE TABLE IF NOT EXISTS ingresos (
            id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
            monto NUMERIC NOT NULL, descripcion TEXT NOT NULL, fecha DATE NOT NULL)""",
        """CREATE TABLE IF NOT EXISTS historial (
            id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
            role TEXT NOT NULL, content TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS presupuestos (
            id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
            categoria TEXT NOT NULL, limite NUMERIC NOT NULL,
            UNIQUE(user_id, categoria))""",
        """CREATE TABLE IF NOT EXISTS metas (
            id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
            nombre TEXT NOT NULL, objetivo NUMERIC NOT NULL, actual NUMERIC DEFAULT 0)""",
        """CREATE TABLE IF NOT EXISTS deudas (
            id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
            tipo TEXT NOT NULL, monto NUMERIC NOT NULL,
            persona TEXT NOT NULL, descripcion TEXT NOT NULL,
            pagada BOOLEAN DEFAULT FALSE, fecha DATE NOT NULL)""",
        """CREATE TABLE IF NOT EXISTS gastos_fijos (
            id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
            descripcion TEXT NOT NULL, monto NUMERIC NOT NULL,
            categoria TEXT NOT NULL, activo BOOLEAN DEFAULT TRUE)""",
        """CREATE TABLE IF NOT EXISTS usuarios (
            user_id BIGINT PRIMARY KEY, nombre TEXT, email TEXT,
            saldo_bolsillo NUMERIC DEFAULT 0, modo_estricto BOOLEAN DEFAULT FALSE,
            last_backup DATE, created_at TIMESTAMP DEFAULT NOW())""",
    ]
    migraciones = [
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS nombre TEXT",
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS saldo_bolsillo NUMERIC DEFAULT 0",
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS modo_estricto BOOLEAN DEFAULT FALSE",
        "ALTER TABLE metas ADD COLUMN IF NOT EXISTS actual NUMERIC DEFAULT 0",
    ]
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            for s in tablas:
                cur.execute(s)
            for s in migraciones:
                try:
                    cur.execute(s)
                except Exception:
                    conn.rollback()
        conn.commit()
    finally:
        release_conn(conn)
    print("DB inicializada.")

# ── Usuarios ──────────────────────────────────────────────────────────────────

def registrar_usuario(user_id, nombre=None):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO usuarios (user_id, nombre) VALUES (%s,%s) ON CONFLICT (user_id) DO NOTHING", (user_id, nombre))
            if nombre:
                cur.execute("UPDATE usuarios SET nombre=%s WHERE user_id=%s AND nombre IS NULL", (nombre, user_id))
        conn.commit()
    finally:
        release_conn(conn)

def get_usuario(user_id):
    return _db_query_one("SELECT * FROM usuarios WHERE user_id=%s", (user_id,), psycopg2.extras.DictCursor)

def get_todos_usuarios():
    rows = _db_query("SELECT user_id FROM usuarios")
    return [r[0] for r in rows]

def actualizar_campo_usuario(user_id, campo, valor):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE usuarios SET {campo}=%s WHERE user_id=%s", (valor, user_id))
        conn.commit()
    finally:
        release_conn(conn)

def toggle_modo_estricto(user_id):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE usuarios SET modo_estricto = NOT modo_estricto WHERE user_id=%s RETURNING modo_estricto", (user_id,))
            r = cur.fetchone()
        conn.commit()
    finally:
        release_conn(conn)
    return bool(r[0]) if r else False

# ── Gastos ────────────────────────────────────────────────────────────────────

def registrar_gasto(user_id, monto, categoria, descripcion):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO gastos (user_id,monto,categoria,descripcion,fecha) VALUES (%s,%s,%s,%s,%s)",
                        (user_id, monto, categoria.strip(), descripcion.strip(), colombia_today()))
        conn.commit()
    finally:
        release_conn(conn)
    u = get_usuario(user_id)
    if u and u["saldo_bolsillo"] and float(u["saldo_bolsillo"]) > 0:
        actualizar_campo_usuario(user_id, "saldo_bolsillo", max(0.0, float(u["saldo_bolsillo"]) - monto))

def borrar_ultimo_gasto(user_id):
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT id,monto,categoria,descripcion FROM gastos WHERE user_id=%s ORDER BY id DESC LIMIT 1", (user_id,))
            g = cur.fetchone()
            if not g:
                return None
            cur.execute("DELETE FROM gastos WHERE id=%s", (g["id"],))
        conn.commit()
        g = dict(g)
    finally:
        release_conn(conn)
    u = get_usuario(user_id)
    if u and u["saldo_bolsillo"] is not None:
        actualizar_campo_usuario(user_id, "saldo_bolsillo", float(u["saldo_bolsillo"]) + float(g["monto"]))
    return g

# ── Ingresos ──────────────────────────────────────────────────────────────────

def registrar_ingreso(user_id, monto, descripcion):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO ingresos (user_id,monto,descripcion,fecha) VALUES (%s,%s,%s,%s)",
                        (user_id, monto, descripcion.strip(), colombia_today()))
        conn.commit()
    finally:
        release_conn(conn)
    u = get_usuario(user_id)
    if u and u["saldo_bolsillo"] is not None:
        actualizar_campo_usuario(user_id, "saldo_bolsillo", float(u["saldo_bolsillo"] or 0) + monto)

# ── Deudas ────────────────────────────────────────────────────────────────────

def registrar_deuda(user_id, tipo, monto, persona, descripcion):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO deudas (user_id,tipo,monto,persona,descripcion,fecha) VALUES (%s,%s,%s,%s,%s,%s)",
                        (user_id, tipo, monto, persona.strip(), descripcion.strip(), colombia_today()))
        conn.commit()
    finally:
        release_conn(conn)

def get_deudas(user_id):
    return _db_query("SELECT * FROM deudas WHERE user_id=%s AND pagada=FALSE ORDER BY fecha DESC",
                     (user_id,), psycopg2.extras.DictCursor)

def marcar_deuda_pagada(user_id, deuda_id):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE deudas SET pagada=TRUE WHERE id=%s AND user_id=%s", (deuda_id, user_id))
        conn.commit()
    finally:
        release_conn(conn)

# ── Gastos fijos ──────────────────────────────────────────────────────────────

def registrar_gasto_fijo(user_id, descripcion, monto, categoria):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO gastos_fijos (user_id,descripcion,monto,categoria) VALUES (%s,%s,%s,%s)",
                        (user_id, descripcion.strip(), monto, categoria.strip()))
        conn.commit()
    finally:
        release_conn(conn)

def get_gastos_fijos(user_id):
    return _db_query("SELECT * FROM gastos_fijos WHERE user_id=%s AND activo=TRUE",
                     (user_id,), psycopg2.extras.DictCursor)

def aplicar_gastos_fijos(user_id):
    fijos = get_gastos_fijos(user_id)
    for f in fijos:
        registrar_gasto(user_id, float(f["monto"]), f["categoria"], f["descripcion"] + " (fijo)")
    return fijos

# ── Presupuestos ──────────────────────────────────────────────────────────────

def guardar_presupuesto(user_id, categoria, limite):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""INSERT INTO presupuestos (user_id,categoria,limite) VALUES (%s,%s,%s)
                           ON CONFLICT (user_id,categoria) DO UPDATE SET limite=EXCLUDED.limite""",
                        (user_id, categoria.strip(), limite))
        conn.commit()
    finally:
        release_conn(conn)

def get_presupuestos(user_id):
    rows = _db_query("SELECT categoria,limite FROM presupuestos WHERE user_id=%s",
                     (user_id,), psycopg2.extras.DictCursor)
    return {r["categoria"]: float(r["limite"]) for r in rows}

def verificar_alerta_presupuesto(user_id, categoria):
    pres = get_presupuestos(user_id)
    if categoria not in pres:
        return None
    limite = pres[categoria]
    mes = colombia_today().replace(day=1)
    row = _db_query_one("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE user_id=%s AND categoria=%s AND fecha>=%s",
                        (user_id, categoria, mes))
    gastado = float(row[0])
    pct = gastado / limite * 100 if limite > 0 else 0
    b = barra(gastado, limite, 10)
    if pct >= 100:
        return f"🚨 Presupuesto SUPERADO en {categoria}\n{b} {pct:.0f}%\nLimite ${limite:,.0f} | Gastado ${gastado:,.0f}"
    if pct >= 80:
        return f"⚠️ Cuidado con {categoria}\n{b} {pct:.0f}%\nRestante ${limite - gastado:,.0f} de ${limite:,.0f}"
    return None

# ── Historial IA ──────────────────────────────────────────────────────────────

def get_historial(user_id, limit=20):
    rows = _db_query(
        "SELECT role,content FROM historial WHERE user_id=%s ORDER BY created_at DESC LIMIT %s",
        (user_id, limit), psycopg2.extras.DictCursor
    )
    return [{"role": r["role"], "content": r["content"]} for r in reversed(rows)]

def guardar_historial(user_id, role, content):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO historial (user_id,role,content) VALUES (%s,%s,%s)", (user_id, role, content))
            cur.execute("DELETE FROM historial WHERE id IN (SELECT id FROM historial WHERE user_id=%s ORDER BY created_at DESC OFFSET 30)", (user_id,))
        conn.commit()
    finally:
        release_conn(conn)

# ── Consultas de datos ────────────────────────────────────────────────────────

def get_resumen_mes_actual(user_id):
    mes = colombia_today().replace(day=1)
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE user_id=%s AND fecha>=%s", (user_id, mes))
            tg = float(cur.fetchone()[0])
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM ingresos WHERE user_id=%s AND fecha>=%s", (user_id, mes))
            ti = float(cur.fetchone()[0])
            cur.execute("SELECT categoria,SUM(monto) AS total FROM gastos WHERE user_id=%s AND fecha>=%s GROUP BY categoria ORDER BY total DESC", (user_id, mes))
            por_cat = {r["categoria"]: float(r["total"]) for r in cur.fetchall()}
    finally:
        release_conn(conn)
    return tg, ti, ti - tg, por_cat

def get_resumen_fecha(user_id, fecha):
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE user_id=%s AND fecha=%s", (user_id, fecha))
            tg = float(cur.fetchone()[0])
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM ingresos WHERE user_id=%s AND fecha=%s", (user_id, fecha))
            ti = float(cur.fetchone()[0])
            cur.execute("SELECT categoria,SUM(monto) AS total FROM gastos WHERE user_id=%s AND fecha=%s GROUP BY categoria ORDER BY total DESC", (user_id, fecha))
            por_cat = {r["categoria"]: float(r["total"]) for r in cur.fetchall()}
    finally:
        release_conn(conn)
    return tg, ti, por_cat

def get_resumen_periodo(user_id, inicio, fin):
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE user_id=%s AND fecha>=%s AND fecha<=%s", (user_id, inicio, fin))
            tg = float(cur.fetchone()[0])
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM ingresos WHERE user_id=%s AND fecha>=%s AND fecha<=%s", (user_id, inicio, fin))
            ti = float(cur.fetchone()[0])
            cur.execute("SELECT categoria,SUM(monto) AS total FROM gastos WHERE user_id=%s AND fecha>=%s AND fecha<=%s GROUP BY categoria ORDER BY total DESC", (user_id, inicio, fin))
            por_cat = {r["categoria"]: float(r["total"]) for r in cur.fetchall()}
    finally:
        release_conn(conn)
    return tg, ti, ti - tg, por_cat

def tuvo_actividad_hoy(user_id):
    hoy = colombia_today()
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM gastos WHERE user_id=%s AND fecha=%s", (user_id, hoy))
            g = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM ingresos WHERE user_id=%s AND fecha=%s", (user_id, hoy))
            i = cur.fetchone()[0]
    finally:
        release_conn(conn)
    return (g + i) > 0

def get_metas(user_id):
    return _db_query("SELECT id,nombre,objetivo,actual FROM metas WHERE user_id=%s",
                     (user_id,), psycopg2.extras.DictCursor)

def registrar_meta(user_id, nombre, objetivo):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO metas (user_id,nombre,objetivo,actual) VALUES (%s,%s,%s,0)",
                        (user_id, nombre.strip(), objetivo))
        conn.commit()
    finally:
        release_conn(conn)

def abonar_meta(user_id, meta_id, monto):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE metas SET actual = LEAST(actual + %s, objetivo) WHERE id=%s AND user_id=%s",
                        (monto, meta_id, user_id))
        conn.commit()
    finally:
        release_conn(conn)

def get_categorias_frecuentes(user_id, limit=5):
    rows = _db_query(
        "SELECT categoria,COUNT(*) AS cnt,SUM(monto) AS total FROM gastos WHERE user_id=%s GROUP BY categoria ORDER BY cnt DESC LIMIT %s",
        (user_id, limit)
    )
    return [(r[0], int(r[1]), float(r[2])) for r in rows]

def get_ultimos_movimientos(user_id, dias=7):
    inicio = colombia_today() - datetime.timedelta(days=dias)
    return _db_query(
        """SELECT fecha,categoria,descripcion,monto,'gasto' AS tipo FROM gastos WHERE user_id=%s AND fecha>=%s
           UNION ALL
           SELECT fecha,'Ingreso' AS categoria,descripcion,monto,'ingreso' AS tipo FROM ingresos WHERE user_id=%s AND fecha>=%s
           ORDER BY fecha DESC LIMIT 15""",
        (user_id, inicio, user_id, inicio), psycopg2.extras.DictCursor
    )

def get_racha_ahorro(user_id):
    hoy = colombia_today()
    mes = hoy.replace(day=1)
    tg, _, _, _ = get_resumen_mes_actual(user_id)
    dias_tx = (hoy - mes).days + 1
    if dias_tx == 0 or tg == 0:
        return 0
    prom = tg / dias_tx
    racha = 0
    for i in range(dias_tx):
        tgd, _, _ = get_resumen_fecha(user_id, hoy - datetime.timedelta(days=i))
        if tgd <= prom:
            racha += 1
        else:
            break
    return racha

def detectar_gasto_inusual(user_id):
    hoy = colombia_today()
    tg_hoy, _, _ = get_resumen_fecha(user_id, hoy)
    if tg_hoy == 0:
        return False, 0, 0
    tg_mes, _, _, _ = get_resumen_mes_actual(user_id)
    dias = (hoy - hoy.replace(day=1)).days + 1
    prom = tg_mes / dias if dias > 0 else 0
    return (prom > 0 and tg_hoy > prom * 2), tg_hoy, prom

# ── Textos de analisis ────────────────────────────────────────────────────────

def texto_comparar_meses(user_id):
    hoy     = colombia_today()
    mes_ant = 12 if hoy.month == 1 else hoy.month - 1
    year_ant = hoy.year - 1 if hoy.month == 1 else hoy.year
    inicio_act = hoy.replace(day=1)
    inicio_ant = datetime.date(year_ant, mes_ant, 1)
    fin_ant    = datetime.date(year_ant, mes_ant, 1).replace(day=28) + datetime.timedelta(days=4)
    fin_ant    = fin_ant - datetime.timedelta(days=fin_ant.day)
    ga, ia, _, ca = get_resumen_periodo(user_id, inicio_act, hoy)
    gb, ib, _, cb = get_resumen_periodo(user_id, inicio_ant, fin_ant)
    meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    t = f"Comparativa {meses[hoy.month-1]} vs {meses[mes_ant-1]}\n{'─'*28}\n\n"
    if gb > 0:
        d = (ga - gb) / gb * 100
        t += f"Gastos\n  Este mes: ${ga:,.0f}\n  Mes ant:  ${gb:,.0f}\n  {'📈' if d>0 else '📉'} {abs(d):.1f}% {'mas' if d>0 else 'menos'}\n\n"
    else:
        t += f"Gastos este mes: ${ga:,.0f} (sin datos anteriores)\n\n"
    if ib > 0:
        d = (ia - ib) / ib * 100
        t += f"Ingresos\n  Este mes: ${ia:,.0f}\n  Mes ant:  ${ib:,.0f}\n  {'📈' if d>0 else '📉'} {abs(d):.1f}% {'mas' if d>0 else 'menos'}\n\n"
    else:
        t += f"Ingresos este mes: ${ia:,.0f}\n\n"
    if cb:
        t += "Cambios por categoria:\n"
        for cat in set(list(ca.keys()) + list(cb.keys())):
            act, ant = ca.get(cat, 0), cb.get(cat, 0)
            if ant > 0:
                d = (act - ant) / ant * 100
                t += f"  {'📈' if d>0 else '📉'} {cat}: {abs(d):.0f}% (${act:,.0f})\n"
            elif act > 0:
                t += f"  🆕 {cat}: ${act:,.0f}\n"
    return t

def texto_prediccion(user_id):
    hoy    = colombia_today()
    inicio = hoy.replace(day=1)
    dias_tx = (hoy - inicio).days + 1
    prox   = datetime.date(hoy.year + 1, 1, 1) if hoy.month == 12 else datetime.date(hoy.year, hoy.month + 1, 1)
    dias_tot = (prox - inicio).days
    tg, ti, _, _ = get_resumen_mes_actual(user_id)
    if dias_tx == 0 or tg == 0:
        return None
    prom = tg / dias_tx
    proy = prom * dias_tot
    rest = dias_tot - dias_tx
    t  = f"Prediccion del mes\n{'─'*22}\n\n"
    t += f"Llevas {dias_tx} de {dias_tot} dias\n"
    t += f"Gastado: ${tg:,.0f}\n"
    t += f"Promedio diario: ${prom:,.0f}\n\n"
    t += f"Proyeccion fin de mes: ${proy:,.0f}\n"
    t += f"Dias restantes: {rest}\n"
    t += f"Gastaras aprox: ${prom * rest:,.0f} mas\n"
    if ti > 0:
        bal_proy = ti - proy
        t += f"\n{'📈' if bal_proy >= 0 else '📉'} Balance proyectado: ${bal_proy:,.0f}"
    return t

def texto_dias_caros(user_id):
    hace_30 = colombia_today() - datetime.timedelta(days=30)
    rows = _db_query(
        "SELECT EXTRACT(DOW FROM fecha) AS dow,SUM(monto) AS total,COUNT(*) AS cnt FROM gastos WHERE user_id=%s AND fecha>=%s GROUP BY dow ORDER BY total DESC",
        (user_id, hace_30)
    )
    if not rows:
        return None
    dias = ["Domingo","Lunes","Martes","Miercoles","Jueves","Viernes","Sabado"]
    max_total = float(rows[0][1])
    t = f"Dias de mayor gasto (30 dias)\n{'─'*28}\n\n"
    for r in rows[:7]:
        d, total, cnt = int(r[0]), float(r[1]), int(r[2])
        t += f"{dias[d]}\n{barra(total, max_total, 8)} ${total:,.0f} ({cnt} gastos)\n\n"
    t += f"Los {dias[int(rows[0][0])]}s son tus dias mas caros."
    return t

# ── Graficas ──────────────────────────────────────────────────────────────────

def _setup_ax(ax):
    ax.set_facecolor("#161b22")
    for s in ["top","right"]: ax.spines[s].set_visible(False)
    for s in ["bottom","left"]: ax.spines[s].set_color("#444")
    ax.tick_params(colors="white")

def generar_grafica_mensual(user_id):
    tg, ti, bal, por_cat = get_resumen_mes_actual(user_id)
    mes = colombia_now().strftime("%B %Y")
    fig = plt.figure(figsize=(10, 14), facecolor="#0d1117")
    fig.suptitle(f"Resumen Financiero — {mes}", fontsize=15, fontweight="bold", color="white", y=0.98)
    if not por_cat and ti == 0:
        ax = fig.add_subplot(111)
        ax.set_facecolor("#0d1117")
        ax.text(0.5, 0.5, "Sin datos este mes", ha="center", va="center", fontsize=14, color="white", transform=ax.transAxes)
        ax.axis("off")
    else:
        gs = fig.add_gridspec(3, 2, hspace=0.55, wspace=0.35)
        ax1 = fig.add_subplot(gs[0, :])
        ax1.set_facecolor("#161b22")
        if por_cat:
            wedges, texts, autos = ax1.pie(por_cat.values(), labels=por_cat.keys(),
                colors=COLORES[:len(por_cat)], autopct="%1.1f%%", startangle=90,
                textprops={"color":"white","fontsize":9}, wedgeprops={"edgecolor":"#0d1117","linewidth":2})
            for a in autos: a.set_color("white"); a.set_fontsize(8)
            ax1.set_title("Gastos por Categoria", color="white", fontsize=12, pad=10)
        ax2 = fig.add_subplot(gs[1, 0])
        _setup_ax(ax2)
        vals = [ti, tg]
        tope = max(vals) if max(vals) > 0 else 1
        bars = ax2.bar(["Ingresos","Gastos"], vals, color=["#06d6a0","#e94560"], width=0.5, edgecolor="#0d1117")
        for b2, v in zip(bars, vals):
            ax2.text(b2.get_x() + b2.get_width()/2, v + tope*0.02, f"${v:,.0f}", ha="center", va="bottom", color="white", fontsize=8, fontweight="bold")
        ax2.set_title("Ingresos vs Gastos", color="white", fontsize=10, pad=8)
        ax3 = fig.add_subplot(gs[1, 1])
        ax3.set_facecolor("#161b22")
        color_bal = "#06d6a0" if bal >= 0 else "#e94560"
        ax3.text(0.5, 0.68, "Balance", ha="center", va="center", fontsize=13, color="white", fontweight="bold", transform=ax3.transAxes)
        ax3.text(0.5, 0.40, f"${abs(bal):,.0f}", ha="center", va="center", fontsize=17, color=color_bal, fontweight="bold", transform=ax3.transAxes)
        ax3.text(0.5, 0.18, "📈 Positivo" if bal >= 0 else "📉 Negativo", ha="center", va="center", fontsize=9, color=color_bal, transform=ax3.transAxes)
        ax3.axis("off")
        if por_cat:
            ax4 = fig.add_subplot(gs[2, :])
            _setup_ax(ax4)
            cats   = list(por_cat.keys())[-8:]
            montos = [por_cat[c] for c in cats]
            y_pos  = list(range(len(cats)))
            tope2  = max(montos) if max(montos) > 0 else 1
            bh = ax4.barh(y_pos, montos, color=COLORES[:len(cats)], edgecolor="#0d1117")
            ax4.set_yticks(y_pos)
            ax4.set_yticklabels(cats, color="white", fontsize=9)
            for b3, v in zip(bh, montos):
                ax4.text(v + tope2*0.01, b3.get_y() + b3.get_height()/2, f"${v:,.0f}", va="center", color="white", fontsize=8)
            ax4.set_title("Detalle por Categoria", color="white", fontsize=10, pad=8)
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=110, bbox_inches="tight", facecolor="#0d1117")
    buf.seek(0)
    plt.close()
    return buf

def generar_grafica_semanal(user_id):
    hoy    = colombia_today()
    fechas = [hoy - datetime.timedelta(days=i) for i in range(6, -1, -1)]
    montos = [get_resumen_fecha(user_id, f)[0] for f in fechas]
    labels = [f.strftime("%d/%m") for f in fechas]
    fig, ax = plt.subplots(figsize=(10, 5), facecolor="#0d1117")
    _setup_ax(ax)
    mx = max(montos) if max(montos) > 0 else 1
    colores_b = ["#e94560" if m == mx else "#06d6a0" for m in montos]
    bars = ax.bar(labels, montos, color=colores_b, edgecolor="#0d1117", width=0.6)
    for b4, v in zip(bars, montos):
        if v > 0:
            ax.text(b4.get_x() + b4.get_width()/2, v + mx*0.02, f"${v:,.0f}", ha="center", va="bottom", color="white", fontsize=8)
    prom = sum(montos) / 7
    if prom > 0:
        ax.axhline(prom, color="#ffd166", linestyle="--", alpha=0.7, linewidth=1.5)
        ax.text(6.4, prom + mx*0.01, f"prom ${prom:,.0f}", color="#ffd166", fontsize=8)
    ax.set_title("Gastos ultimos 7 dias", color="white", fontsize=13, fontweight="bold", pad=12)
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=110, bbox_inches="tight", facecolor="#0d1117")
    buf.seek(0)
    plt.close()
    return buf

# ── Excel ─────────────────────────────────────────────────────────────────────

def generar_excel(user_id):
    mes       = colombia_today().replace(day=1)
    mes_nombre = colombia_now().strftime("%B %Y")
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT fecha,categoria,descripcion,monto FROM gastos WHERE user_id=%s AND fecha>=%s ORDER BY fecha", (user_id, mes))
            gastos = cur.fetchall()
            cur.execute("SELECT fecha,descripcion,monto FROM ingresos WHERE user_id=%s AND fecha>=%s ORDER BY fecha", (user_id, mes))
            ingresos = cur.fetchall()
    finally:
        release_conn(conn)
    wb     = openpyxl.Workbook()
    H_FILL = PatternFill("solid", fgColor="1a1a2e")
    H_FONT = Font(bold=True, color="FFFFFF")
    ALT    = PatternFill("solid", fgColor="F2F2F2")

    ws1 = wb.active; ws1.title = "Gastos"
    for col, h in enumerate(["Fecha","Categoria","Descripcion","Monto"], 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal="center")
    for i, w in zip(range(1,5), [14,18,30,16]):
        ws1.column_dimensions[chr(64+i)].width = w
    tg = 0.0
    for i, g in enumerate(gastos, 2):
        ws1.cell(row=i, column=1, value=str(g["fecha"]))
        ws1.cell(row=i, column=2, value=g["categoria"])
        ws1.cell(row=i, column=3, value=g["descripcion"])
        c = ws1.cell(row=i, column=4, value=float(g["monto"])); c.number_format = "$#,##0"
        tg += float(g["monto"])
        if i % 2 == 0:
            for col in range(1,5): ws1.cell(row=i, column=col).fill = ALT
    ft = len(gastos) + 2
    ws1.cell(row=ft, column=3, value="TOTAL").font = Font(bold=True)
    c = ws1.cell(row=ft, column=4, value=tg); c.font = Font(bold=True, color="E94560"); c.number_format = "$#,##0"

    ws2 = wb.create_sheet("Ingresos")
    for col, h in enumerate(["Fecha","Descripcion","Monto"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal="center")
    for i, w in zip(range(1,4), [14,30,16]):
        ws2.column_dimensions[chr(64+i)].width = w
    ti = 0.0
    for i, ing in enumerate(ingresos, 2):
        ws2.cell(row=i, column=1, value=str(ing["fecha"]))
        ws2.cell(row=i, column=2, value=ing["descripcion"])
        c = ws2.cell(row=i, column=3, value=float(ing["monto"])); c.number_format = "$#,##0"
        ti += float(ing["monto"])
        if i % 2 == 0:
            for col in range(1,4): ws2.cell(row=i, column=col).fill = ALT
    fi = len(ingresos) + 2
    ws2.cell(row=fi, column=2, value="TOTAL").font = Font(bold=True)
    c = ws2.cell(row=fi, column=3, value=ti); c.font = Font(bold=True, color="06D6A0"); c.number_format = "$#,##0"

    ws3 = wb.create_sheet("Resumen")
    ws3.column_dimensions["A"].width = 22; ws3.column_dimensions["B"].width = 18
    ws3["A1"] = f"Resumen — {mes_nombre}"; ws3["A1"].font = Font(bold=True, size=14)
    bal = ti - tg
    for row, (lbl, val) in enumerate([("Total Ingresos", ti), ("Total Gastos", tg), ("Balance", bal)], 3):
        ws3.cell(row=row, column=1, value=lbl).font = Font(bold=True)
        v = ws3.cell(row=row, column=2, value=val); v.number_format = "$#,##0"
        if lbl == "Balance":
            v.font = Font(bold=True, color="06D6A0" if bal >= 0 else "E94560")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, mes_nombre

# ── IA ────────────────────────────────────────────────────────────────────────

def _llamar_ia(messages, max_tokens=600):
    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        json={"model": "claude-haiku-4-5-20251001", "max_tokens": max_tokens, "system": SYSTEM_PROMPT, "messages": messages},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["content"][0]["text"]

async def llamar_ia_async(messages, max_tokens=600):
    """Llama a la IA sin bloquear el event loop de Telegram."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, lambda: _llamar_ia(messages, max_tokens))

async def pedir_consejo_ia(user_id):
    tg, ti, bal, pc = get_resumen_mes_actual(user_id)
    racha = get_racha_ahorro(user_id)
    inusual, tg_hoy, prom = detectar_gasto_inusual(user_id)
    metas = get_metas(user_id)
    metas_txt = ""
    if metas:
        m = metas[0]
        act = float(m["actual"] or 0)
        obj = float(m["objetivo"])
        pct = act / obj * 100 if obj > 0 else 0
        metas_txt = f", meta '{m['nombre']}' al {pct:.0f}%"
    ctx = (f"Finanzas del usuario: ingresos=${ti:,.0f}, gastos=${tg:,.0f}, balance=${bal:,.0f}, "
           f"categorias_top={dict(list(pc.items())[:3])}, racha_ahorro={racha} dias, "
           f"gasto_inusual_hoy={inusual}{metas_txt}. "
           f"Da un consejo financiero personalizado (3-4 oraciones), motivador y especifico. "
           f"Menciona numeros concretos de sus datos. Sin markdown.")
    try:
        return await llamar_ia_async([{"role":"user","content":ctx}], max_tokens=250)
    except Exception:
        return "Registra tus movimientos regularmente para recibir consejos personalizados 📊"

async def analizar_recibo(photo_bytes):
    img_b64 = base64.b64encode(photo_bytes).decode("utf-8")
    def _call():
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={"model":"claude-haiku-4-5-20251001","max_tokens":200,
                  "messages":[{"role":"user","content":[
                      {"type":"image","source":{"type":"base64","media_type":"image/jpeg","data":img_b64}},
                      {"type":"text","text":"Analiza este recibo. Extrae el TOTAL y descripcion. Responde SOLO: MONTO|DESCRIPCION|CATEGORIA. Ejemplo: 25000|Almuerzo|Comida. Si no se lee: ERROR|no_legible"},
                  ]}]},
            timeout=60,
        )
        return resp.json()["content"][0]["text"]
    try:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, _call)
    except Exception as e:
        print(f"Error recibo: {e}")
        return None

# ── Email ─────────────────────────────────────────────────────────────────────

def enviar_email_backup(destinatario, archivo_bytes, mes_nombre, user_id):
    if not EMAIL_SENDER or not EMAIL_PASSWORD:
        return False, "Credenciales no configuradas"
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_SENDER; msg["To"] = destinatario
        msg["Subject"] = f"FinBot — Backup {mes_nombre}"
        msg.attach(MIMEText(f"Hola!\n\nAqui esta tu backup de FinBot de {mes_nombre}.\n\nSaludos,\nFinBot", "plain"))
        archivo_bytes.seek(0)
        adj = MIMEBase("application","octet-stream"); adj.set_payload(archivo_bytes.read())
        encoders.encode_base64(adj)
        adj.add_header("Content-Disposition", f"attachment; filename=FinBot_{mes_nombre.replace(' ','_')}.xlsx")
        msg.attach(adj)
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls(); server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.send_message(msg); server.quit()
        actualizar_campo_usuario(user_id, "last_backup", colombia_today())
        return True, "OK"
    except Exception as e:
        print(f"Error email: {e}"); return False, str(e)

# ── Scheduler ─────────────────────────────────────────────────────────────────

async def enviar_buenos_dias(bot):
    hoy = colombia_today()
    tip = TIPS[hoy.day % len(TIPS)]
    for user_id in get_todos_usuarios():
        try:
            u      = get_usuario(user_id)
            nombre = u["nombre"] if u and u["nombre"] else ""
            t      = f"Buenos dias{', ' + nombre if nombre else ''}! 🌅\n\n{tip}"
            metas  = get_metas(user_id)
            if metas:
                m   = metas[0]
                act = float(m["actual"] or 0)
                obj = float(m["objetivo"])
                pct = act / obj * 100 if obj > 0 else 0
                t  += f"\n\nMeta: {m['nombre']}\n{barra(act, obj, 10)} {pct:.0f}%\n${act:,.0f} / ${obj:,.0f}"
            await bot.send_message(chat_id=user_id, text=t)
        except Exception as e:
            print(f"Error buenos dias {user_id}: {e}")

async def enviar_resumenes_noche(bot):
    hoy    = colombia_today()
    es_dom = hoy.weekday() == 6
    for user_id in get_todos_usuarios():
        try:
            if es_dom:
                inicio = hoy - datetime.timedelta(days=6)
                tgs, tis, bal, pcs = get_resumen_periodo(user_id, inicio, hoy)
                t = f"Resumen semanal 📅\n{'─'*22}\n"
                t += f"Ingresos: ${tis:,.0f}\nGastos:   ${tgs:,.0f}\nBalance:  ${bal:,.0f}\n"
                if pcs:
                    t += "\nPor categoria:\n"
                    for c, m in sorted(pcs.items(), key=lambda x: x[1], reverse=True):
                        t += f"  {c}: ${m:,.0f}\n"
                racha = get_racha_ahorro(user_id)
                if racha >= 3:
                    t += f"\n🔥 Racha: {racha} dias sin pasarte del promedio!"
                await bot.send_message(chat_id=user_id, text=t)
                try:
                    await bot.send_photo(chat_id=user_id, photo=generar_grafica_semanal(user_id), caption="Tu semana en grafica 📊")
                except Exception:
                    pass

            if tuvo_actividad_hoy(user_id):
                tg, ti, por_cat = get_resumen_fecha(user_id, hoy)
                t  = f"Resumen de hoy {hoy.strftime('%d/%m/%Y')} 📊\n{'─'*22}\n"
                if ti > 0: t += f"Ingresos: ${ti:,.0f}\n"
                if tg > 0:
                    t += f"Gastos:   ${tg:,.0f}\n"
                    for c, m in sorted(por_cat.items(), key=lambda x: x[1], reverse=True):
                        t += f"  {c}: ${m:,.0f}\n"
                bd = ti - tg
                t += f"\n{'📈' if bd >= 0 else '📉'} Balance: ${bd:,.0f}"
                inusual, tg_hoy, prom = detectar_gasto_inusual(user_id)
                if inusual:
                    t += f"\n\n⚠️ Hoy gastaste ${tg_hoy:,.0f}, mas del doble de tu promedio (${prom:,.0f})"
                await bot.send_message(chat_id=user_id, text=t)
            else:
                await bot.send_message(chat_id=user_id, text="No registraste nada hoy 🔔\nGastaste o ganaste algo? Cuentame 📊")
        except Exception as e:
            print(f"Error resumen noche {user_id}: {e}")

async def enviar_backups_mensuales(bot):
    rows = _db_query("SELECT user_id,email FROM usuarios WHERE email IS NOT NULL AND email != ''")
    for user_id, email in rows:
        try:
            buf, mes_nombre = generar_excel(user_id)
            ok, _ = enviar_email_backup(email, buf, mes_nombre, user_id)
            if ok:
                await bot.send_message(chat_id=user_id, text=f"Backup del mes enviado a {email} 📧")
        except Exception as e:
            print(f"Error backup {user_id}: {e}")

async def aplicar_fijos_todos(bot):
    for user_id in get_todos_usuarios():
        try:
            fijos = aplicar_gastos_fijos(user_id)
            if fijos:
                total = sum(float(f["monto"]) for f in fijos)
                t = "Gastos fijos del mes aplicados 🔁\n"
                for f in fijos:
                    t += f"  {f['descripcion']}: ${float(f['monto']):,.0f}\n"
                t += f"\nTotal: ${total:,.0f}"
                await bot.send_message(chat_id=user_id, text=t)
        except Exception as e:
            print(f"Error fijos {user_id}: {e}")

async def scheduler(bot):
    """Scheduler robusto: se reinicia si falla, nunca muere silenciosamente."""
    enviado_manana = None
    enviado_noche  = None
    backup_mes     = None
    print("Scheduler iniciado.")
    while True:
        try:
            ahora = colombia_now()
            hoy   = ahora.date()
            hora  = ahora.hour
            min_  = ahora.minute

            if hora == 8 and min_ < 10 and enviado_manana != hoy:
                enviado_manana = hoy
                print(f"Enviando buenos dias {hoy}")
                await enviar_buenos_dias(bot)

            if hora == 20 and min_ < 10 and enviado_noche != hoy:
                enviado_noche = hoy
                print(f"Enviando resumenes noche {hoy}")
                await enviar_resumenes_noche(bot)

            if hoy.day == 1 and hora == 9 and min_ < 10 and backup_mes != hoy.month:
                backup_mes = hoy.month
                print(f"Enviando backups mensuales {hoy}")
                await enviar_backups_mensuales(bot)
                await aplicar_fijos_todos(bot)

        except Exception as e:
            print(f"Error en scheduler (continuando): {e}")

        await asyncio.sleep(55)

# ── Web server ────────────────────────────────────────────────────────────────

class HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200); self.send_header("Content-type","text/plain"); self.end_headers()
        self.wfile.write(b"OK")
    def log_message(self, fmt, *args):
        pass

def run_web_server():
    HTTPServer(("0.0.0.0", int(os.environ.get("PORT", 10000))), HealthHandler).serve_forever()

# ── UI helpers ────────────────────────────────────────────────────────────────

def kb_principal():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💸 Gasto",               callback_data="menu_gasto"),
         InlineKeyboardButton("💰 Ingreso",             callback_data="menu_ingreso")],
        [InlineKeyboardButton("📊 Resumen",             callback_data="menu_resumen"),
         InlineKeyboardButton("📈 Grafica mensual",     callback_data="menu_grafica")],
        [InlineKeyboardButton("📊 Grafica semanal",     callback_data="menu_grafica_sem"),
         InlineKeyboardButton("🔮 Prediccion",          callback_data="menu_prediccion")],
        [InlineKeyboardButton("📅 Comparar meses",      callback_data="menu_comparar"),
         InlineKeyboardButton("📆 Dias caros",          callback_data="menu_dias")],
        [InlineKeyboardButton("📋 Ultimos movimientos", callback_data="menu_ultimos"),
         InlineKeyboardButton("🎯 Metas",               callback_data="menu_metas")],
        [InlineKeyboardButton("🔔 Presupuestos",        callback_data="menu_presupuestos"),
         InlineKeyboardButton("🏦 Deudas",              callback_data="menu_deudas")],
        [InlineKeyboardButton("🔁 Gastos fijos",        callback_data="menu_fijos"),
         InlineKeyboardButton("💵 Mi saldo",            callback_data="menu_saldo")],
        [InlineKeyboardButton("😤 Modo estricto",       callback_data="menu_estricto"),
         InlineKeyboardButton("💡 Consejo IA",          callback_data="menu_consejo")],
        [InlineKeyboardButton("📥 Excel",               callback_data="menu_excel"),
         InlineKeyboardButton("📧 Mi correo",           callback_data="menu_email")],
    ])

async def mostrar_resumen(message, user_id):
    tg, ti, bal, pc = get_resumen_mes_actual(user_id)
    u   = get_usuario(user_id)
    mes = colombia_now().strftime("%B %Y")
    t   = f"Resumen {mes} 📊\n{'─'*24}\n\n"
    t  += f"Ingresos:  ${ti:,.0f}\n"
    t  += f"Gastos:    ${tg:,.0f}\n"
    t  += f"{'📈' if bal >= 0 else '📉'} Balance:  ${bal:,.0f}\n"
    if u and u["saldo_bolsillo"] and float(u["saldo_bolsillo"]) > 0:
        t += f"\n💵 Bolsillo: ${float(u['saldo_bolsillo']):,.0f}\n"
    if pc:
        t += "\nPor categoria:\n"
        pres = get_presupuestos(user_id)
        for c, m in sorted(pc.items(), key=lambda x: x[1], reverse=True):
            if c in pres:
                pct = m / pres[c] * 100 if pres[c] > 0 else 0
                t += f"  {c}: ${m:,.0f}\n  {barra(m, pres[c], 8)} {pct:.0f}%\n"
            else:
                t += f"  {c}: ${m:,.0f}\n"
    racha = get_racha_ahorro(user_id)
    if racha >= 2:
        t += f"\n🔥 Racha: {racha} dias en control!"
    if not pc and ti == 0:
        t += "\nAun no hay registros este mes."
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("📈 Mensual",  callback_data="menu_grafica"),
        InlineKeyboardButton("📊 Semanal",  callback_data="menu_grafica_sem"),
    ],[
        InlineKeyboardButton("📥 Excel",    callback_data="menu_excel"),
        InlineKeyboardButton("💡 Consejo",  callback_data="menu_consejo"),
    ]])
    await message.reply_text(t, reply_markup=kb)

# ── Tokens ────────────────────────────────────────────────────────────────────

TOKEN_RE = {
    "borrar":  re.compile(r'BORRAR_ULTIMO_GASTO'),
    "gasto":   re.compile(r'GASTO_REGISTRADO\|([^|\n]+)\|([^|\n]+)\|([^\n]+)'),
    "ingreso": re.compile(r'INGRESO_REGISTRADO\|([^|\n]+)\|([^\n]+)'),
    "presup":  re.compile(r'PRESUPUESTO_DEFINIDO\|([^|\n]+)\|([^\n]+)'),
    "email":   re.compile(r'EMAIL_CONFIGURADO\|([^\s\n]+)'),
    "saldo":   re.compile(r'SALDO_ACTUALIZADO\|([^\n]+)'),
    "deuda":   re.compile(r'DEUDA_REGISTRADA\|([^|\n]+)\|([^|\n]+)\|([^|\n]+)\|([^\n]+)'),
    "fijo":    re.compile(r'GASTO_FIJO_REGISTRADO\|([^|\n]+)\|([^|\n]+)\|([^\n]+)'),
    "meta":    re.compile(r'META_REGISTRADA\|([^|\n]+)\|([^\n]+)'),
}

def _comentario_ia(reply):
    """Extrae el texto de la IA despues del token."""
    limpio = re.sub(r'[A-Z_]+(?:\|[^\n]*)?\n?', '', reply, count=1).strip()
    return _limpiar_markdown(limpio)

async def procesar_respuesta_ia(reply, user_id, update):
    u = get_usuario(user_id)

    if TOKEN_RE["borrar"].search(reply):
        g = borrar_ultimo_gasto(user_id)
        if g:
            t = f"Borrado el ultimo gasto ↩️\n💸 ${float(g['monto']):,.0f} — {g['descripcion']} ({g['categoria']})"
        else:
            t = "No encontre ningun gasto reciente para borrar."
        await update.message.reply_text(t)
        guardar_historial(user_id, "assistant", reply)
        return True

    m = TOKEN_RE["gasto"].search(reply)
    if m:
        try:
            monto = _parse_monto(m.group(1))
            cat   = m.group(2).strip()
            desc  = m.group(3).strip()
            registrar_gasto(user_id, monto, cat, desc)
            u = get_usuario(user_id)  # refrescar tras actualizar saldo
            saldo_txt = ""
            if u and u["saldo_bolsillo"] and float(u["saldo_bolsillo"]) > 0:
                saldo_txt = f"\n💵 Bolsillo: ${float(u['saldo_bolsillo']):,.0f}"
            cmt = _comentario_ia(reply)
            t   = f"Gasto registrado ✅\n💸 ${monto:,.0f} — {desc}\n📂 {cat}{saldo_txt}"
            t  += f"\n\n{cmt if cmt else frase_gasto(monto)}"
            await update.message.reply_text(t)
            alerta = verificar_alerta_presupuesto(user_id, cat)
            if alerta:
                await update.message.reply_text(alerta)
            if u and u["modo_estricto"]:
                inusual, tg_hoy, prom = detectar_gasto_inusual(user_id)
                if inusual:
                    await update.message.reply_text(f"😤 Modo estricto: Hoy gastaste ${tg_hoy:,.0f}. Tu promedio es ${prom:,.0f}. Controla!")
            guardar_historial(user_id, "assistant", reply)
            return True
        except Exception as e:
            print(f"Error parsing gasto: {e}")

    m = TOKEN_RE["ingreso"].search(reply)
    if m:
        try:
            monto = _parse_monto(m.group(1))
            desc  = m.group(2).strip()
            registrar_ingreso(user_id, monto, desc)
            cmt = _comentario_ia(reply)
            t   = f"Ingreso registrado ✅\n💰 ${monto:,.0f} — {desc}"
            t  += f"\n\n{cmt if cmt else frase_ingreso(monto)}"
            await update.message.reply_text(t)
            guardar_historial(user_id, "assistant", reply)
            return True
        except Exception as e:
            print(f"Error parsing ingreso: {e}")

    m = TOKEN_RE["presup"].search(reply)
    if m:
        try:
            cat    = m.group(1).strip()
            limite = _parse_monto(m.group(2))
            guardar_presupuesto(user_id, cat, limite)
            await update.message.reply_text(f"Presupuesto guardado ✅\n📂 {cat}: ${limite:,.0f}/mes\nTe aviso cuando llegues al 80% 🔔")
            guardar_historial(user_id, "assistant", reply)
            return True
        except Exception as e:
            print(f"Error parsing presupuesto: {e}")

    m = TOKEN_RE["email"].search(reply)
    if m:
        email = m.group(1).strip()
        if "@" in email and "." in email:
            actualizar_campo_usuario(user_id, "email", email)
            await update.message.reply_text(f"Correo guardado: {email} 📧\nRecibiras tu backup el dia 1 de cada mes.")
            guardar_historial(user_id, "assistant", reply)
            return True

    m = TOKEN_RE["saldo"].search(reply)
    if m:
        try:
            saldo = _parse_monto(m.group(1))
            actualizar_campo_usuario(user_id, "saldo_bolsillo", saldo)
            await update.message.reply_text(f"Saldo actualizado: ${saldo:,.0f} 💵\nAhora llevo el control de tu bolsillo.")
            guardar_historial(user_id, "assistant", reply)
            return True
        except Exception as e:
            print(f"Error parsing saldo: {e}")

    m = TOKEN_RE["deuda"].search(reply)
    if m:
        try:
            tipo    = m.group(1).strip()
            monto   = _parse_monto(m.group(2))
            persona = m.group(3).strip()
            desc    = m.group(4).strip()
            registrar_deuda(user_id, tipo, monto, persona, desc)
            emoji  = "💸" if tipo == "debo" else "💰"
            tipo_t = "Deuda registrada" if tipo == "debo" else "Credito registrado"
            await update.message.reply_text(f"{tipo_t} {emoji}\n{persona}: ${monto:,.0f}\n{desc}")
            guardar_historial(user_id, "assistant", reply)
            return True
        except Exception as e:
            print(f"Error parsing deuda: {e}")

    m = TOKEN_RE["fijo"].search(reply)
    if m:
        try:
            desc  = m.group(1).strip()
            monto = _parse_monto(m.group(2))
            cat   = m.group(3).strip()
            registrar_gasto_fijo(user_id, desc, monto, cat)
            await update.message.reply_text(f"Gasto fijo registrado 🔁\n{desc}: ${monto:,.0f}/mes ({cat})\nSe aplica automaticamente el dia 1.")
            guardar_historial(user_id, "assistant", reply)
            return True
        except Exception as e:
            print(f"Error parsing fijo: {e}")

    m = TOKEN_RE["meta"].search(reply)
    if m:
        try:
            nombre  = m.group(1).strip()
            objetivo = _parse_monto(m.group(2))
            registrar_meta(user_id, nombre, objetivo)
            cmt = _comentario_ia(reply)
            await update.message.reply_text(
                f"Meta creada 🎯\n{nombre}: ${objetivo:,.0f}\n\n{cmt if cmt else 'Puedes abonar diciendome: abone 50000 a ' + nombre}"
            )
            guardar_historial(user_id, "assistant", reply)
            return True
        except Exception as e:
            print(f"Error parsing meta: {e}")

    return False

# ── Handlers de Telegram ──────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    nombre  = update.effective_user.first_name or ""
    registrar_usuario(user_id, nombre)
    saludo = f"Hola, {nombre}! 👋" if nombre else "Hola! 👋"
    await update.message.reply_text(
        f"{saludo} Soy FinBot, tu asistente financiero 🤖\n\n"
        "Escribeme en lenguaje natural:\n"
        "  💸 gaste 50000 en comida\n"
        "  💰 gane 2000000\n"
        "  💵 mi saldo es 300000\n"
        "  🏦 debo 80000 a Juan\n"
        "  📸 Foto de recibo = registro automatico\n"
        "  ↩️ me equivoque = borra el ultimo gasto\n\n"
        "O usa los botones:",
        reply_markup=kb_principal(),
    )

async def cmd_resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await mostrar_resumen(update.message, update.effective_user.id)

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q       = update.callback_query
    await q.answer()
    user_id = q.from_user.id
    data    = q.data

    if data == "menu_gasto":
        cats = get_categorias_frecuentes(user_id)
        t = 'Escribe: "gaste 50000 en comida"\n\n📸 Tambien puedes enviar foto del recibo!'
        if cats:
            t += "\n\nTus categorias frecuentes:\n"
            for c, n, total in cats:
                t += f"  {c} ({n}x — ${total:,.0f} total)\n"
        await q.message.reply_text(t)

    elif data == "menu_ingreso":
        await q.message.reply_text('Escribe: "gane 2000000" o "recibi 500000 por freelance"')

    elif data == "menu_resumen":
        await mostrar_resumen(q.message, user_id)

    elif data == "menu_grafica":
        msg = await q.message.reply_text("Generando grafica mensual...")
        await q.message.reply_photo(photo=generar_grafica_mensual(user_id), caption="Tu resumen del mes 📊")
        await msg.delete()

    elif data == "menu_grafica_sem":
        msg = await q.message.reply_text("Generando grafica semanal...")
        await q.message.reply_photo(photo=generar_grafica_semanal(user_id), caption="Tus ultimos 7 dias 📊")
        await msg.delete()

    elif data == "menu_prediccion":
        p = texto_prediccion(user_id)
        await q.message.reply_text(p if p else "Registra mas gastos para poder hacer predicciones 😊")

    elif data == "menu_comparar":
        await q.message.reply_text(texto_comparar_meses(user_id))

    elif data == "menu_dias":
        d = texto_dias_caros(user_id)
        await q.message.reply_text(d if d else "Necesito al menos una semana de datos para este analisis.")

    elif data == "menu_ultimos":
        movs = get_ultimos_movimientos(user_id)
        if not movs:
            await q.message.reply_text("No hay movimientos en los ultimos 7 dias.")
            return
        t = f"Ultimos movimientos (7 dias) 📋\n{'─'*26}\n\n"
        for m in movs:
            emoji = "💸" if m["tipo"] == "gasto" else "💰"
            t += f"{emoji} {m['fecha'].strftime('%d/%m')} — {m['descripcion']}\n"
            t += f"   {m['categoria']} | ${float(m['monto']):,.0f}\n\n"
        await q.message.reply_text(t)

    elif data == "menu_metas":
        metas = get_metas(user_id)
        if not metas:
            await q.message.reply_text('Sin metas aun.\n\nEscribe: "meta ahorrar 500000 para vacaciones"')
            return
        t = "Tus metas 🎯\n\n"
        for m in metas:
            act = float(m["actual"] or 0)
            obj = float(m["objetivo"])
            pct = act / obj * 100 if obj > 0 else 0
            t  += f"{m['nombre']}\n{barra(act, obj, 12)} {pct:.0f}%\n${act:,.0f} / ${obj:,.0f}\n\n"
        await q.message.reply_text(t)

    elif data == "menu_presupuestos":
        pres = get_presupuestos(user_id)
        if not pres:
            await q.message.reply_text('Sin presupuestos.\n\nEscribe: "presupuesto comida 300000"')
            return
        _, _, _, pc = get_resumen_mes_actual(user_id)
        t = "Tus presupuestos 🔔\n\n"
        for c, l in pres.items():
            g   = pc.get(c, 0)
            pct = g / l * 100 if l > 0 else 0
            e   = "🚨" if pct >= 100 else "⚠️" if pct >= 80 else "✅"
            t  += f"{e} {c}\n{barra(g, l, 10)} {pct:.0f}%\n${g:,.0f} / ${l:,.0f}\n\n"
        await q.message.reply_text(t)

    elif data == "menu_deudas":
        deudas = get_deudas(user_id)
        if not deudas:
            await q.message.reply_text('Sin deudas activas 🙌\n\nEscribe:\n"debo 50000 a Juan por el almuerzo"\n"me deben 100000 Pedro"')
            return
        debo     = [d for d in deudas if d["tipo"] == "debo"]
        me_deben = [d for d in deudas if d["tipo"] == "medeben"]
        t = "Deudas activas 🏦\n\n"
        if debo:
            t += f"Lo que DEBES (${sum(float(d['monto']) for d in debo):,.0f}):\n"
            for d in debo:
                t += f"  💸 {d['persona']}: ${float(d['monto']):,.0f} — {d['descripcion']}\n"
            t += "\n"
        if me_deben:
            t += f"Lo que TE DEBEN (${sum(float(d['monto']) for d in me_deben):,.0f}):\n"
            for d in me_deben:
                t += f"  💰 {d['persona']}: ${float(d['monto']):,.0f} — {d['descripcion']}\n"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("✅ Marcar como pagada", callback_data="menu_pagar_deuda")]])
        await q.message.reply_text(t, reply_markup=kb)

    elif data == "menu_pagar_deuda":
        deudas = get_deudas(user_id)
        if not deudas:
            await q.message.reply_text("No tienes deudas pendientes.")
            return
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton(
                f"{'Debo' if d['tipo']=='debo' else 'Me deben'} ${float(d['monto']):,.0f} — {d['persona']}",
                callback_data=f"pagar_{d['id']}",
            )
        ] for d in deudas])
        await q.message.reply_text("Selecciona la deuda saldada:", reply_markup=kb)

    elif data.startswith("pagar_"):
        marcar_deuda_pagada(user_id, int(data.split("_")[1]))
        await q.message.reply_text("Deuda marcada como pagada! 🎉")

    elif data == "menu_fijos":
        fijos = get_gastos_fijos(user_id)
        if not fijos:
            await q.message.reply_text('Sin gastos fijos.\n\nEscribe: "gasto fijo arriendo cada mes 800000"')
            return
        total = sum(float(f["monto"]) for f in fijos)
        t = "Gastos fijos mensuales 🔁\n\n"
        for f in fijos:
            t += f"  {f['descripcion']} ({f['categoria']}): ${float(f['monto']):,.0f}\n"
        t += f"\nTotal mensual: ${total:,.0f}"
        await q.message.reply_text(t)

    elif data == "menu_saldo":
        u     = get_usuario(user_id)
        saldo = float(u["saldo_bolsillo"] or 0) if u else 0
        await q.message.reply_text(f"Tu saldo en bolsillo: ${saldo:,.0f} 💵\n\nPara actualizar: \"mi saldo es 300000\"")

    elif data == "menu_estricto":
        nuevo = toggle_modo_estricto(user_id)
        if nuevo:
            await q.message.reply_text("Modo estricto ACTIVADO 😤\nCada vez que te pases del promedio te lo digo. Vamos!")
        else:
            await q.message.reply_text("Modo estricto desactivado 😌\nVolvemos a modo amable.")

    elif data == "menu_consejo":
        msg = await q.message.reply_text("Analizando tus finanzas...")
        c   = await pedir_consejo_ia(user_id)
        await msg.edit_text(f"Consejo personalizado 💡\n{'─'*22}\n\n{c}")

    elif data == "menu_excel":
        msg = await q.message.reply_text("Generando Excel...")
        buf, mn = generar_excel(user_id)
        await q.message.reply_document(document=buf, filename=f"FinBot_{mn.replace(' ','_')}.xlsx", caption=f"Tu Excel de {mn} 📊")
        await msg.delete()

    elif data == "menu_email":
        u     = get_usuario(user_id)
        email = u["email"] if u else None
        if email:
            await q.message.reply_text(f"Tu correo: {email}\nBackup automatico cada dia 1.\n\nPara cambiarlo: \"mi correo es nuevo@email.com\"")
        else:
            await q.message.reply_text('Sin correo.\n\nEscribe: "mi correo es tucorreo@gmail.com"')

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    registrar_usuario(user_id, update.effective_user.first_name or "")
    msg = await update.message.reply_text("Analizando recibo 📸...")
    try:
        photo_file  = await update.message.photo[-1].get_file()
        photo_bytes = await photo_file.download_as_bytearray()
        resultado   = await analizar_recibo(bytes(photo_bytes))
        if not resultado or resultado.startswith("ERROR"):
            await msg.edit_text("No pude leer el recibo. Intenta con mejor iluminacion o registralo manualmente.")
            return
        parts = resultado.strip().split("|")
        if len(parts) < 3:
            await msg.edit_text("No reconoci el formato. Registralo manualmente.")
            return
        monto = _parse_monto(parts[0])
        desc  = parts[1].strip()
        cat   = parts[2].strip()
        registrar_gasto(user_id, monto, cat, desc)
        await msg.edit_text(f"Gasto del recibo registrado!\n💸 ${monto:,.0f}\n📂 {cat}\n📝 {desc}")
        alerta = verificar_alerta_presupuesto(user_id, cat)
        if alerta:
            await update.message.reply_text(alerta)
    except Exception as e:
        print(f"Error photo: {e}")
        await msg.edit_text("Error procesando la foto. Intenta de nuevo.")

async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Notas de voz en desarrollo 🎤\nEscribeme el gasto o envia foto del recibo 📸")

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id   = update.effective_user.id
    nombre    = update.effective_user.first_name or ""
    user_text = update.message.text.strip()
    registrar_usuario(user_id, nombre)

    tl = user_text.lower()
    if tl in ("/resumen","resumen","balance"):
        await mostrar_resumen(update.message, user_id); return
    if tl in ("/grafica","grafica"):
        await update.message.reply_photo(photo=generar_grafica_mensual(user_id), caption="Tu resumen del mes 📊"); return
    if tl in ("/excel","excel"):
        buf, mn = generar_excel(user_id)
        await update.message.reply_document(document=buf, filename=f"FinBot_{mn.replace(' ','_')}.xlsx"); return
    if tl in ("/prediccion","prediccion"):
        p = texto_prediccion(user_id)
        await update.message.reply_text(p if p else "Registra mas gastos para predicciones."); return
    if tl in ("/comparar","comparar"):
        await update.message.reply_text(texto_comparar_meses(user_id)); return

    # Comando para abonar a meta: "abone 50000 a vacaciones"
    m_abono = re.match(r'(?:abone?|abonar?|agrega?r?)\s+([\d.,]+)\s+(?:a|en)\s+(.+)', tl)
    if m_abono:
        try:
            monto_abono = _parse_monto(m_abono.group(1))
            nombre_meta = m_abono.group(2).strip().title()
            metas = get_metas(user_id)
            meta_match = next((m for m in metas if nombre_meta.lower() in m["nombre"].lower()), None)
            if meta_match:
                abonar_meta(user_id, meta_match["id"], monto_abono)
                metas = get_metas(user_id)
                meta_upd = next(m for m in metas if m["id"] == meta_match["id"])
                act = float(meta_upd["actual"] or 0)
                obj = float(meta_upd["objetivo"])
                pct = act / obj * 100 if obj > 0 else 0
                await update.message.reply_text(
                    f"Abono registrado a {meta_upd['nombre']} ✅\n"
                    f"{barra(act, obj, 12)} {pct:.0f}%\n"
                    f"${act:,.0f} / ${obj:,.0f}"
                    + ("\n\nMeta completada! 🎉" if pct >= 100 else f"\nFaltan ${obj - act:,.0f}")
                )
                return
        except Exception as e:
            print(f"Error abono meta: {e}")

    historial = get_historial(user_id)
    historial.append({"role":"user","content":user_text})
    guardar_historial(user_id, "user", user_text)
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")

    try:
        reply = await llamar_ia_async(historial)
    except Exception as e:
        print(f"Error API: {e}")
        await update.message.reply_text("Error de conexion, intenta de nuevo 🙏")
        return

    procesado = await procesar_respuesta_ia(reply, user_id, update)
    if not procesado:
        guardar_historial(user_id, "assistant", reply)
        await update.message.reply_text(_limpiar_markdown(reply))

# ── Main ──────────────────────────────────────────────────────────────────────

async def post_init(application):
    async def _scheduler_guardado():
        """Reinicia el scheduler si muere inesperadamente."""
        while True:
            try:
                await scheduler(application.bot)
            except Exception as e:
                print(f"Scheduler murio, reiniciando en 30s: {e}")
                await asyncio.sleep(30)
    asyncio.create_task(_scheduler_guardado())

def main():
    init_db()
    threading.Thread(target=run_web_server, daemon=True).start()
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).post_init(post_init).build()
    app.add_handler(CommandHandler("start",   cmd_start))
    app.add_handler(CommandHandler("resumen", cmd_resumen))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.PHOTO,                   handle_photo))
    app.add_handler(MessageHandler(filters.VOICE,                   handle_voice))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    print("FinBot v3 listo!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
